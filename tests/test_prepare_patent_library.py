from __future__ import annotations

import csv
import codecs
from hashlib import sha256
import os
from pathlib import Path
import stat
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from tools.prepare_patent_library import prepare_patent_library
import tools.prepare_patent_library as preparation_module


SOURCE_HEADERS = [
    "抗体名称",
    "类型(IgG/VHH)",
    "抗体重链氨基酸",
    "抗体轻链氨基酸(若有)",
    "是否结合PVRIG",
    "是否阻断PVRL2",
    "来自专利",
    "开发公司",
    "相关药物管线",
]


def write_source(path: Path) -> None:
    workbook = Workbook()
    library = workbook.active
    library.title = "参考阳参抗体"
    library.append(SOURCE_HEADERS)
    library.append(["Repeat", "VHH", " acd ef ", None, "是", "是", "P1", "Co", "Drug"])
    library.append(["Repeat", "VHH", "GHIK", None, "否", "是", "P2", "Co", "Drug"])
    library.append(["FirstIgG", "IgG", "LMNP Q", "RST VW", "是", "否", "P3", "Co", "Drug"])
    library.append(["DuplicateIgG", "IgG", "LMNPQ", "RSTVW", "是", "否", "P4", "Co", "Drug"])
    library.append(["备注：仅供核对", None, None, None, None, None, None, None, None])
    library["A6"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    submission = workbook.create_sheet("抗体提交格式")
    submission.append(
        [
            "序号",
            "抗体名称",
            "重链VH可变区",
            "轻链VL可变区（若为VHH，此处为空）",
            "推荐排序",
            "优化改造/从头设计",
            "起始分子重链序列",
            "起始分子轻链序列（若为VHH，此处为空）",
        ]
    )
    submission.append([1, "existing", "ACDEF", None, 1, "从头设计", None, None])
    workbook.save(path)


def run_preparation(tmp_path: Path, **kwargs):
    source = tmp_path / "source.xlsx"
    write_source(source)
    return prepare_patent_library(
        source_path=source,
        cleaned_path=tmp_path / "cleaned.xlsx",
        csv_path=tmp_path / "positive.csv",
        benchmark_path=tmp_path / "benchmark.xlsx",
        benchmark_size=2,
        **kwargs,
    )


def set_default_column_alignment(path: Path, columns_by_sheet: dict[str, str]) -> None:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    workbook_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationship_namespace = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_relationship_namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    with ZipFile(path) as archive:
        files = {name: archive.read(name) for name in archive.namelist()}

    styles = ET.fromstring(files["xl/styles.xml"])
    cell_xfs = styles.find(f"{{{namespace}}}cellXfs")
    assert cell_xfs is not None
    default_xf = cell_xfs[0]
    default_xf.set("applyAlignment", "1")
    ET.SubElement(default_xf, f"{{{namespace}}}alignment", {"vertical": "center"})
    files["xl/styles.xml"] = ET.tostring(styles, encoding="utf-8", xml_declaration=True)

    workbook = ET.fromstring(files["xl/workbook.xml"])
    relationships = ET.fromstring(files["xl/_rels/workbook.xml.rels"])
    targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships.findall(f"{{{package_relationship_namespace}}}Relationship")
    }
    sheet_paths = {
        sheet.attrib["name"]: targets[sheet.attrib[f"{{{relationship_namespace}}}id"]].lstrip("/")
        for sheet in workbook.findall(f"{{{workbook_namespace}}}sheets/{{{workbook_namespace}}}sheet")
    }
    for sheet_name, column in columns_by_sheet.items():
        worksheet = ET.fromstring(files[sheet_paths[sheet_name]])
        columns = worksheet.find(f"{{{namespace}}}cols")
        assert columns is not None
        column_index = ord(column) - ord("A") + 1
        dimension = next(
            item
            for item in columns.findall(f"{{{namespace}}}col")
            if item.attrib["min"] == str(column_index) and item.attrib["max"] == str(column_index)
        )
        dimension.set("style", "0")
        files[sheet_paths[sheet_name]] = ET.tostring(worksheet, encoding="utf-8", xml_declaration=True)

    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)


def set_empty_core_properties(path: Path) -> None:
    with ZipFile(path) as archive:
        files = {name: archive.read(name) for name in archive.namelist()}
    files["docProps/core.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        b'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"/>'
    )
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)


def core_property_tags(path: Path) -> list[str]:
    with ZipFile(path) as archive:
        root = ET.fromstring(archive.read("docProps/core.xml"))
    return [element.tag for element in root]


def test_prepare_patent_library_cleans_exports_and_preserves_workbook(tmp_path):
    summary = run_preparation(tmp_path)

    assert summary.source_records == 4
    assert summary.cleaned_records == 3
    assert summary.igg_records == 1
    assert summary.vhh_records == 2
    assert summary.renamed_records == 1
    assert summary.removed_duplicate_records == 1
    assert summary.source_sha256 == sha256((tmp_path / "source.xlsx").read_bytes()).hexdigest()

    cleaned = load_workbook(tmp_path / "cleaned.xlsx")
    assert cleaned.sheetnames == ["参考阳参抗体", "抗体提交格式"]
    rows = list(cleaned.worksheets[0].iter_rows(values_only=True))
    assert [row[0] for row in rows[1:]] == ["Repeat", "Repeat-1", "FirstIgG", "备注：仅供核对"]
    assert rows[1][2] == "ACDEF"
    assert rows[3][2:4] == ("LMNPQ", "RSTVW")
    assert cleaned.worksheets[0]["A5"].fill.fgColor.rgb == "0000FF00"

    with (tmp_path / "positive.csv").open(encoding="utf-8-sig", newline="") as file:
        csv_rows = list(csv.reader(file))
    assert (tmp_path / "positive.csv").read_bytes().startswith(codecs.BOM_UTF8)
    assert csv_rows[0] == SOURCE_HEADERS
    assert csv_rows[1][0:4] == ["Repeat", "VHH", "ACDEF", ""]
    assert csv_rows[2][0:4] == ["Repeat-1", "VHH", "GHIK", ""]
    assert csv_rows[3][0:4] == ["FirstIgG", "IgG", "LMNPQ", "RSTVW"]
    assert len(csv_rows) == 4
    assert all(len(row) == 9 for row in csv_rows)

    benchmark = load_workbook(tmp_path / "benchmark.xlsx")
    benchmark_rows = list(benchmark.active.iter_rows(values_only=True))
    assert len(benchmark.worksheets) == 1
    assert [row[1] for row in benchmark_rows[1:]] == ["benchmark_001_Repeat", "benchmark_002_Repeat-1"]
    assert [row[2] for row in benchmark_rows[1:]] == ["ACDEF", "GHIK"]
    assert all(row[3] is None and row[6] is None and row[7] is None for row in benchmark_rows[1:])


def test_prepare_patent_library_exports_csv_with_lf_line_endings(tmp_path):
    run_preparation(tmp_path)

    csv_bytes = (tmp_path / "positive.csv").read_bytes()

    assert b"\r\n" not in csv_bytes
    assert csv_bytes.count(b"\n") == 4


def test_prepare_patent_library_rejects_unexpected_source_hash(tmp_path):
    with pytest.raises(ValueError, match="SHA-256"):
        run_preparation(tmp_path, expected_sha256="not-the-source-hash")


def test_prepare_patent_library_requires_requested_vhh_count(tmp_path):
    source = tmp_path / "source.xlsx"
    write_source(source)

    with pytest.raises(ValueError, match="VHH"):
        prepare_patent_library(
            source_path=source,
            cleaned_path=tmp_path / "cleaned.xlsx",
            csv_path=tmp_path / "positive.csv",
            benchmark_path=tmp_path / "benchmark.xlsx",
            benchmark_size=3,
        )


def test_prepare_patent_library_preserves_later_unique_name_during_renaming(tmp_path):
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    library = workbook.active
    library.title = "参考阳参抗体"
    library.append(SOURCE_HEADERS)
    library.append(["A", "VHH", "ACD", None, None, None, None, None, None])
    library.append(["A", "VHH", "EFG", None, None, None, None, None, None])
    library.append(["A-1", "VHH", "HIK", None, None, None, None, None, None])
    workbook.create_sheet("抗体提交格式")
    workbook.save(source)

    summary = prepare_patent_library(
        source_path=source,
        cleaned_path=tmp_path / "cleaned.xlsx",
        csv_path=tmp_path / "positive.csv",
        benchmark_path=tmp_path / "benchmark.xlsx",
        benchmark_size=3,
    )

    cleaned = load_workbook(tmp_path / "cleaned.xlsx")
    names = [row[0] for row in cleaned.worksheets[0].iter_rows(min_row=2, values_only=True)]
    assert names == ["A", "A-2", "A-1"]
    assert names[2] == "A-1"
    assert len(names) == len(set(names))
    assert summary.renamed_records == 1


def test_prepare_patent_library_preserves_hyperlinks_and_row_heights_after_deletion(tmp_path):
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    library = workbook.active
    library.title = "参考阳参抗体"
    library.append(SOURCE_HEADERS)
    library.append(["First", "VHH", "ACD", None, None, None, None, None, None])
    library.append(["Duplicate", "VHH", "ACD", None, None, None, None, None, None])
    library.append(["Linked", "VHH", "EFG", None, None, None, None, None, None])
    library["A4"].hyperlink = "https://example.test/linked"
    library.row_dimensions[4].height = 33
    workbook.create_sheet("抗体提交格式")
    workbook.save(source)

    prepare_patent_library(
        source_path=source,
        cleaned_path=tmp_path / "cleaned.xlsx",
        csv_path=tmp_path / "positive.csv",
        benchmark_path=tmp_path / "benchmark.xlsx",
        benchmark_size=2,
    )

    cleaned = load_workbook(tmp_path / "cleaned.xlsx")
    linked = cleaned.worksheets[0]["A3"]
    assert linked.value == "Linked"
    assert linked.hyperlink.target == "https://example.test/linked"
    assert linked.hyperlink.location is None
    assert cleaned.worksheets[0].row_dimensions[3].height == 33


def test_prepare_patent_library_preserves_column_alignment_on_each_sheet(tmp_path):
    source = tmp_path / "source.xlsx"
    write_source(source)
    workbook = load_workbook(source)
    workbook["参考阳参抗体"].column_dimensions["A"].width = 22
    workbook["抗体提交格式"].column_dimensions["B"].width = 18
    workbook.save(source)
    set_default_column_alignment(source, {"参考阳参抗体": "A", "抗体提交格式": "B"})

    prepare_patent_library(
        source_path=source,
        cleaned_path=tmp_path / "cleaned.xlsx",
        csv_path=tmp_path / "positive.csv",
        benchmark_path=tmp_path / "benchmark.xlsx",
        benchmark_size=2,
    )

    cleaned = load_workbook(tmp_path / "cleaned.xlsx")
    assert cleaned["参考阳参抗体"].column_dimensions["A"].alignment.vertical == "center"
    assert cleaned["抗体提交格式"].column_dimensions["B"].alignment.vertical == "center"


def test_prepare_patent_library_preserves_empty_core_properties(tmp_path):
    source = tmp_path / "source.xlsx"
    write_source(source)
    set_empty_core_properties(source)

    prepare_patent_library(
        source_path=source,
        cleaned_path=tmp_path / "cleaned.xlsx",
        csv_path=tmp_path / "positive.csv",
        benchmark_path=tmp_path / "benchmark.xlsx",
        benchmark_size=2,
    )

    assert core_property_tags(tmp_path / "cleaned.xlsx") == core_property_tags(source) == []


def test_prepare_patent_library_keeps_existing_outputs_when_validation_fails(tmp_path):
    source = tmp_path / "source.xlsx"
    write_source(source)
    targets = {
        "cleaned": tmp_path / "cleaned.xlsx",
        "csv": tmp_path / "positive.csv",
        "benchmark": tmp_path / "benchmark.xlsx",
    }
    before = {name: f"old-{name}".encode() for name, path in targets.items()}
    for name, path in targets.items():
        path.write_bytes(before[name])

    with pytest.raises(ValueError, match="VHH"):
        prepare_patent_library(
            source_path=source,
            cleaned_path=targets["cleaned"],
            csv_path=targets["csv"],
            benchmark_path=targets["benchmark"],
            benchmark_size=3,
        )

    assert {name: path.read_bytes() for name, path in targets.items()} == before


@pytest.mark.parametrize("conflict", ["cleaned", "csv", "benchmark", "outputs"])
def test_prepare_patent_library_rejects_conflicting_paths_without_changing_source(tmp_path, conflict):
    source = tmp_path / "source.xlsx"
    write_source(source)
    cleaned = tmp_path / "cleaned.xlsx"
    csv_path = tmp_path / "positive.csv"
    benchmark = tmp_path / "benchmark.xlsx"
    if conflict == "cleaned":
        cleaned = source
    elif conflict == "csv":
        csv_path = source
    elif conflict == "benchmark":
        benchmark = source
    else:
        csv_path = cleaned
    source_bytes = source.read_bytes()

    with pytest.raises(ValueError, match="paths must be different"):
        prepare_patent_library(
            source_path=source,
            cleaned_path=cleaned,
            csv_path=csv_path,
            benchmark_path=benchmark,
            benchmark_size=2,
        )

    assert source.read_bytes() == source_bytes


@pytest.mark.parametrize(
    "row, message",
    [
        (["Bad", "Other", "ACD", None, None, None, None, None, None], "row 2"),
        (["Missing", "VHH", None, None, None, None, None, None, None], "row 2"),
    ],
)
def test_prepare_patent_library_rejects_invalid_data_rows(tmp_path, row, message):
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(SOURCE_HEADERS)
    worksheet.append(row)
    workbook.create_sheet("抗体提交格式")
    workbook.save(source)

    with pytest.raises(ValueError, match=message):
        prepare_patent_library(
            source_path=source,
            cleaned_path=tmp_path / "cleaned.xlsx",
            csv_path=tmp_path / "positive.csv",
            benchmark_path=tmp_path / "benchmark.xlsx",
            benchmark_size=0,
        )


def _existing_targets(tmp_path: Path) -> tuple[dict[str, Path], dict[str, bytes]]:
    targets = {
        "cleaned": tmp_path / "cleaned.xlsx",
        "csv": tmp_path / "positive.csv",
        "benchmark": tmp_path / "benchmark.xlsx",
    }
    before = {name: f"old-{name}".encode() for name in targets}
    for name, path in targets.items():
        path.write_bytes(before[name])
    return targets, before


def test_prepare_patent_library_rolls_back_published_outputs_after_publish_failure(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    write_source(source)
    targets, before = _existing_targets(tmp_path)
    original_replace = preparation_module.os.replace

    def fail_second_publish(source_path, target_path):
        if Path(target_path) == targets["csv"] and not str(source_path).endswith(".backup"):
            raise OSError("publication failure")
        return original_replace(source_path, target_path)

    monkeypatch.setattr(preparation_module.os, "replace", fail_second_publish)

    with pytest.raises(RuntimeError, match="publication failure"):
        prepare_patent_library(
            source_path=source,
            cleaned_path=targets["cleaned"],
            csv_path=targets["csv"],
            benchmark_path=targets["benchmark"],
            benchmark_size=2,
        )

    assert {name: path.read_bytes() for name, path in targets.items()} == before


def test_prepare_patent_library_preserves_backup_when_rollback_fails(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    write_source(source)
    targets, before = _existing_targets(tmp_path)
    original_replace = preparation_module.os.replace

    def fail_publish_then_rollback(source_path, target_path):
        source_path = Path(source_path)
        target_path = Path(target_path)
        if target_path == targets["csv"] and source_path.suffix != ".backup":
            raise OSError("publication failure")
        if target_path == targets["cleaned"] and source_path.suffix == ".backup":
            raise OSError("rollback failure")
        return original_replace(source_path, target_path)

    monkeypatch.setattr(preparation_module.os, "replace", fail_publish_then_rollback)

    with pytest.raises(RuntimeError, match="publication failure.*rollback failure") as error:
        prepare_patent_library(
            source_path=source,
            cleaned_path=targets["cleaned"],
            csv_path=targets["csv"],
            benchmark_path=targets["benchmark"],
            benchmark_size=2,
        )

    backups = list(tmp_path.glob(".cleaned.xlsx.*.backup"))
    assert len(backups) == 1
    assert str(backups[0].resolve()) in str(error.value)
    assert backups[0].read_bytes() == before["cleaned"]
    assert targets["csv"].read_bytes() == before["csv"]
    assert targets["benchmark"].read_bytes() == before["benchmark"]


def test_prepare_patent_library_preserves_existing_output_permissions(tmp_path):
    source = tmp_path / "source.xlsx"
    write_source(source)
    targets, _ = _existing_targets(tmp_path)
    for path in targets.values():
        path.chmod(0o644)

    prepare_patent_library(
        source_path=source,
        cleaned_path=targets["cleaned"],
        csv_path=targets["csv"],
        benchmark_path=targets["benchmark"],
        benchmark_size=2,
    )

    assert all(stat.S_IMODE(path.stat().st_mode) == 0o644 for path in targets.values())


def test_prepare_patent_library_uses_default_permissions_for_new_outputs(tmp_path):
    source = tmp_path / "source.xlsx"
    write_source(source)
    previous_umask = os.umask(0)
    os.umask(previous_umask)
    expected_mode = 0o666 & ~previous_umask

    run_preparation(tmp_path)

    assert stat.S_IMODE((tmp_path / "cleaned.xlsx").stat().st_mode) == expected_mode
