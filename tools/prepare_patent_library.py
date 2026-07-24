"""Prepare the patent antibody library and a VHH benchmark submission workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import shutil
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook


SOURCE_HEADERS = (
    "抗体名称",
    "类型(IgG/VHH)",
    "抗体重链氨基酸",
    "抗体轻链氨基酸(若有)",
    "是否结合PVRIG",
    "是否阻断PVRL2",
    "来自专利",
    "开发公司",
    "相关药物管线",
)
BENCHMARK_HEADERS = (
    "序号",
    "抗体名称",
    "重链VH可变区",
    "轻链VL可变区（若为VHH，此处为空）",
    "推荐排序",
    "优化改造/从头设计",
    "起始分子重链序列",
    "起始分子轻链序列（若为VHH，此处为空）",
)
AMINO_ACIDS = frozenset("ACDEFGHIKLMNPQRSTVWY")
SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(frozen=True)
class PreparationSummary:
    source_records: int
    cleaned_records: int
    igg_records: int
    vhh_records: int
    renamed_records: int
    removed_duplicate_records: int
    source_sha256: str


@dataclass
class _Record:
    row: int
    name: str
    kind: str
    vh: str
    vl: str | None
    metadata: list[object | None]


def prepare_patent_library(
    *,
    source_path: Path,
    cleaned_path: Path,
    csv_path: Path,
    benchmark_path: Path,
    benchmark_size: int = 50,
    expected_sha256: str | None = None,
) -> PreparationSummary:
    """Clean a patent source workbook and export its positive and VHH libraries."""
    source_path, cleaned_path, csv_path, benchmark_path = _validate_paths(
        source_path, cleaned_path, csv_path, benchmark_path
    )
    source_sha256 = hashlib.sha256(source_path.read_bytes()).hexdigest()
    if expected_sha256 is not None and source_sha256 != expected_sha256:
        raise ValueError("source SHA-256 does not match expected_sha256")
    if benchmark_size < 0:
        raise ValueError("benchmark_size must not be negative")

    workbook = load_workbook(source_path)
    worksheet = workbook.worksheets[0]
    records = _read_records(worksheet)
    kept, removed_rows = _remove_sequence_duplicates(records)
    renamed_records = _rename_records(kept)
    _ensure_unique(kept)
    vhh_records = [record for record in kept if record.kind == "VHH"]
    if len(vhh_records) < benchmark_size:
        raise ValueError(f"requested {benchmark_size} VHH records, found {len(vhh_records)}")

    for record in kept:
        worksheet.cell(record.row, 1).value = record.name
        worksheet.cell(record.row, 3).value = record.vh
        worksheet.cell(record.row, 4).value = record.vl
    _delete_rows_preserving_links_and_dimensions(worksheet, removed_rows)
    _write_outputs_transactionally(
        source_path=source_path,
        workbook=workbook,
        cleaned_path=cleaned_path,
        csv_path=csv_path,
        benchmark_path=benchmark_path,
        records=kept,
        benchmark_records=vhh_records[:benchmark_size],
    )

    return PreparationSummary(
        source_records=len(records),
        cleaned_records=len(kept),
        igg_records=sum(record.kind == "IgG" for record in kept),
        vhh_records=len(vhh_records),
        renamed_records=renamed_records,
        removed_duplicate_records=len(removed_rows),
        source_sha256=source_sha256,
    )


def _validate_paths(
    source_path: Path, cleaned_path: Path, csv_path: Path, benchmark_path: Path
) -> tuple[Path, Path, Path, Path]:
    paths = tuple(Path(path).resolve() for path in (source_path, cleaned_path, csv_path, benchmark_path))
    if len(set(paths)) != len(paths):
        raise ValueError("source, cleaned, csv, and benchmark paths must be different")
    return paths


def _read_records(worksheet) -> list[_Record]:
    records = []
    for row in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row, column).value for column in range(1, worksheet.max_column + 1)]
        if not _row_has_data(values):
            continue
        name = _cell_text(worksheet.cell(row, 1).value)
        kind = _cell_text(worksheet.cell(row, 2).value)
        vh_value = worksheet.cell(row, 3).value
        vl_raw = _cell_text(worksheet.cell(row, 4).value)
        if name.startswith("备注") and not kind and not _cell_text(vh_value) and not vl_raw:
            continue
        if kind not in {"IgG", "VHH"}:
            raise ValueError(f"row {row}: type must be IgG or VHH")
        if not _cell_text(vh_value):
            raise ValueError(f"row {row}: VH is required")
        if not name:
            raise ValueError(f"row {row}: antibody name is required")
        vh = _normalise_sequence(vh_value, row, "VH")
        if kind == "IgG" and not vl_raw:
            raise ValueError(f"row {row}: IgG requires VL")
        if kind == "VHH" and vl_raw:
            raise ValueError(f"row {row}: VHH must not have VL")
        vl = _normalise_sequence(vl_raw, row, "VL") if vl_raw else None
        records.append(
            _Record(
                row=row,
                name=name,
                kind=kind,
                vh=vh,
                vl=vl,
                metadata=[worksheet.cell(row, column).value for column in range(1, 10)],
            )
        )
    return records


def _row_has_data(values: list[object | None]) -> bool:
    return any(value is not None and (not isinstance(value, str) or value.strip()) for value in values)


def _cell_text(value: object | None) -> str:
    return str(value).strip() if value is not None else ""


def _normalise_sequence(value: object, row: int, label: str) -> str:
    sequence = re.sub(r"\s+", "", str(value)).upper()
    if not sequence or not set(sequence) <= AMINO_ACIDS:
        raise ValueError(f"row {row}: {label} contains non-standard amino acids")
    return sequence


def _remove_sequence_duplicates(records: list[_Record]) -> tuple[list[_Record], list[int]]:
    seen: set[tuple[str, str | None]] = set()
    kept: list[_Record] = []
    removed_rows: list[int] = []
    for record in records:
        key = (record.vh, record.vl)
        if key in seen:
            removed_rows.append(record.row)
        else:
            seen.add(key)
            kept.append(record)
    return kept, removed_rows


def _rename_records(records: list[_Record]) -> int:
    original_names = {record.name for record in records}
    occurrences: dict[str, int] = {}
    used: set[str] = set()
    renamed = 0
    for record in records:
        original = record.name
        occurrences[original] = occurrences.get(original, 0) + 1
        if occurrences[original] == 1:
            candidate = original
        else:
            suffix = 1
            candidate = f"{original}-{suffix}"
            while candidate in original_names or candidate in used:
                suffix += 1
                candidate = f"{original}-{suffix}"
        record.name = candidate
        used.add(candidate)
        renamed += candidate != original
    return renamed


def _ensure_unique(records: list[_Record]) -> None:
    names = [record.name for record in records]
    sequences = [(record.vh, record.vl) for record in records]
    if len(names) != len(set(names)):
        raise ValueError("cleaned antibody names are not unique")
    if len(sequences) != len(set(sequences)):
        raise ValueError("cleaned VH/VL pairs are not unique")


def _delete_rows_preserving_links_and_dimensions(worksheet, removed_rows: list[int]) -> None:
    if not removed_rows:
        return
    removed = set(removed_rows)
    hyperlinks = {
        (cell.row, cell.column): copy(cell.hyperlink)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.row not in removed and cell.hyperlink is not None
    }
    dimensions = {
        row: copy(dimension)
        for row, dimension in worksheet.row_dimensions.items()
        if row not in removed
    }

    for row in sorted(removed, reverse=True):
        worksheet.delete_rows(row)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.hyperlink = None
    for row in list(worksheet.row_dimensions):
        del worksheet.row_dimensions[row]

    def new_row(old_row: int) -> int:
        return old_row - sum(deleted_row < old_row for deleted_row in removed)

    for (old_row, column), hyperlink in hyperlinks.items():
        worksheet.cell(new_row(old_row), column).hyperlink = hyperlink
    for old_row, dimension in dimensions.items():
        dimension.index = new_row(old_row)
        worksheet.row_dimensions[dimension.index] = dimension


def _write_csv(path: Path, records: list[_Record]) -> None:
    with Path(path).open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, lineterminator="\n")
        writer.writerow(SOURCE_HEADERS)
        for record in records:
            metadata = record.metadata.copy()
            metadata[0] = record.name
            metadata[1] = record.kind
            metadata[2] = record.vh
            metadata[3] = record.vl
            writer.writerow(metadata)


def _write_benchmark(path: Path, records: list[_Record]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "参赛选手提交"
    worksheet.append(BENCHMARK_HEADERS)
    for index, record in enumerate(records, start=1):
        worksheet.append(
            [
                index,
                f"benchmark_{index:03d}_{record.name}",
                record.vh,
                None,
                index,
                "从头设计",
                None,
                None,
            ]
        )
    workbook.save(path)


def _restore_source_workbook_format(source_path: Path, output_path: Path) -> None:
    """Restore OOXML parts that openpyxl cannot round-trip without changing them."""
    with ZipFile(source_path) as source_archive, ZipFile(output_path) as output_archive:
        source_columns = _columns_by_sheet_title(source_archive)
        output_paths = _worksheet_paths_by_title(output_archive)
        output_files = {name: output_archive.read(name) for name in output_archive.namelist()}
        if "docProps/core.xml" in source_archive.namelist():
            output_files["docProps/core.xml"] = source_archive.read("docProps/core.xml")

    for title, output_sheet_path in output_paths.items():
        if title not in source_columns:
            continue
        worksheet = ET.fromstring(output_files[output_sheet_path])
        columns = worksheet.find(f"{{{SPREADSHEET_NAMESPACE}}}cols")
        source_columns_element = ET.fromstring(source_columns[title])
        if columns is None:
            sheet_data = worksheet.find(f"{{{SPREADSHEET_NAMESPACE}}}sheetData")
            worksheet.insert(list(worksheet).index(sheet_data), source_columns_element)
        else:
            worksheet.insert(list(worksheet).index(columns), source_columns_element)
            worksheet.remove(columns)
        output_files[output_sheet_path] = ET.tostring(worksheet, encoding="utf-8", xml_declaration=True)

    descriptor, replacement_name = tempfile.mkstemp(
        dir=output_path.parent, prefix=f".{output_path.name}.", suffix=output_path.suffix
    )
    os.close(descriptor)
    replacement_path = Path(replacement_name)
    try:
        with ZipFile(replacement_path, "w", ZIP_DEFLATED) as archive:
            for name, content in output_files.items():
                archive.writestr(name, content)
        os.chmod(replacement_path, _file_mode(output_path))
        os.replace(replacement_path, output_path)
    finally:
        replacement_path.unlink(missing_ok=True)


def _columns_by_sheet_title(archive: ZipFile) -> dict[str, bytes]:
    paths = _worksheet_paths_by_title(archive)
    columns = {}
    for title, path in paths.items():
        worksheet = ET.fromstring(archive.read(path))
        element = worksheet.find(f"{{{SPREADSHEET_NAMESPACE}}}cols")
        if element is not None:
            columns[title] = ET.tostring(element, encoding="utf-8")
    return columns


def _worksheet_paths_by_title(archive: ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships.findall(f"{{{PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationship")
    }
    return {
        sheet.attrib["name"]: _worksheet_member_path(targets[sheet.attrib[f"{{{RELATIONSHIP_NAMESPACE}}}id"]])
        for sheet in workbook.findall(f"{{{SPREADSHEET_NAMESPACE}}}sheets/{{{SPREADSHEET_NAMESPACE}}}sheet")
    }


def _worksheet_member_path(target: str) -> str:
    member_path = target.lstrip("/")
    return member_path if member_path.startswith("xl/") else f"xl/{member_path}"


def _write_outputs_transactionally(
    *, source_path: Path, workbook, cleaned_path: Path, csv_path: Path, benchmark_path: Path,
    records: list[_Record], benchmark_records: list[_Record]
) -> None:
    targets = (cleaned_path, csv_path, benchmark_path)
    temporary_paths: list[Path] = []
    backup_paths: dict[Path, Path] = {}
    published: list[Path] = []
    restored_backups: set[Path] = set()
    retained_backups: set[Path] = set()
    publish_completed = False
    try:
        default_mode = _default_file_mode()
        target_modes = {
            target: _file_mode(target) if target.exists() else default_mode for target in targets
        }
        for target in targets:
            descriptor, temporary_name = tempfile.mkstemp(
                dir=target.parent, prefix=f".{target.name}.", suffix=target.suffix
            )
            os.close(descriptor)
            temporary_path = Path(temporary_name)
            os.chmod(temporary_path, target_modes[target])
            temporary_paths.append(temporary_path)
        workbook.save(temporary_paths[0])
        _restore_source_workbook_format(source_path, temporary_paths[0])
        _write_csv(temporary_paths[1], records)
        _write_benchmark(temporary_paths[2], benchmark_records)

        for target in targets:
            if target.exists():
                descriptor, backup_name = tempfile.mkstemp(
                    dir=target.parent, prefix=f".{target.name}.", suffix=".backup"
                )
                os.close(descriptor)
                backup = Path(backup_name)
                shutil.copy2(target, backup)
                backup_paths[target] = backup
        for target, temporary_path in zip(targets, temporary_paths):
            os.replace(temporary_path, target)
            published.append(target)
        publish_completed = True
    except Exception as error:
        rollback_errors: list[tuple[Path, Path | None, OSError]] = []
        for target in reversed(published):
            backup = backup_paths.get(target)
            try:
                if backup is not None:
                    os.replace(backup, target)
                    restored_backups.add(backup)
                else:
                    target.unlink(missing_ok=True)
            except OSError as rollback_error:
                if backup is not None:
                    retained_backups.add(backup)
                rollback_errors.append((target, backup, rollback_error))
        if rollback_errors:
            details = "; ".join(
                f"target {target.resolve()}: rollback error {rollback_error}; "
                f"backup retained at {backup.resolve() if backup is not None else 'none'}"
                for target, backup, rollback_error in rollback_errors
            )
            raise RuntimeError(
                "failed to publish prepared patent library outputs: "
                f"publication error: {error}; {details}"
            ) from error
        raise RuntimeError(
            "failed to publish prepared patent library outputs: "
            f"publication error: {error}"
        ) from error
    finally:
        for path in temporary_paths:
            path.unlink(missing_ok=True)
        for target, path in backup_paths.items():
            if (
                path not in retained_backups
                and (publish_completed or path in restored_backups or target not in published)
            ):
                path.unlink(missing_ok=True)


def _file_mode(path: Path) -> int:
    return path.stat().st_mode & 0o777


def _default_file_mode() -> int:
    current_umask = os.umask(0)
    os.umask(current_umask)
    return 0o666 & ~current_umask


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--cleaned", required=True, type=Path)
    parser.add_argument("--csv", required=True, type=Path)
    parser.add_argument("--benchmark", required=True, type=Path)
    parser.add_argument("--benchmark-size", type=int, default=50)
    parser.add_argument("--expected-sha256")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    summary = prepare_patent_library(
        source_path=args.source,
        cleaned_path=args.cleaned,
        csv_path=args.csv,
        benchmark_path=args.benchmark,
        benchmark_size=args.benchmark_size,
        expected_sha256=args.expected_sha256,
    )
    print(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
